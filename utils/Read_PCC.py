import sys
import time
import openpyxl
import os,shutil
import easygui
import re
import numpy as np



def files_reader(filepath):
    workbook = openpyxl.load_workbook(filepath, data_only=True)
    workbook_formula = openpyxl.load_workbook(filepath, data_only=False)
    sheet_names = []
    for name in workbook.sheetnames:
        if 'PCC' in name:
            sheet_names.append(name)
    worksheet = workbook[sheet_names[0]]   # 读取Sheet
    worksheet_formula = workbook_formula[sheet_names[0]]
    rows, cols = worksheet.max_row, worksheet.max_column
    return rows, cols, worksheet, worksheet_formula

def write2excel(path,data):
    workbook = openpyxl.load_workbook(path)
    worksheet = workbook.worksheets[0]

    for i in data:
        worksheet.append(i)
    workbook.save(path)


def process(writefile_path, rows, cols, worksheet):

    data=[4011, None, None, 1, 0, None, '4011000X', "PP01", 100000, 0, 0, 0, 0]
    part_num = [7, 4]
    po_value = []
    po_index = []


    data[1] = worksheet.cell(part_num[0], part_num[1]).value
    for i in [12, 13, 14]:
        v = worksheet.cell(1, i).value
        if v is None:
            continue
        ret = re.search(r'[A-Z]\d+', str(v)[-3:])
        if ret is not None:
            data[2] = ret.group()
            break
    data[3] = int(data[2][1:]) + 1
    f_info = [data[1], data[2], data[3]]

    for i in range(1, rows+1):
        row = worksheet.cell(row=i, column=1).value
        if row == 'S No.':
            start_no = i
        if type(row) == int and row >0:
            # print(row)
            po_value.append(row)
            po_index.append(i)
    # print(po_value,po_index)

    # BOM_List
    po_value_BOM = po_value.copy()
    po_index_BOM = po_index.copy()
    star_BOM = po_value_BOM.index(11)
    try:
        end_BOM = po_value_BOM.index(12)
    except:
        start_index = star_BOM + 1
        start = po_value_BOM[start_index]
        while True:
            if start + 1 != po_value_BOM[start_index + 1]:
                end_BOM = start_index + 1
                break
            start += 1
            start_index += 1

    po_index_BOM = po_index_BOM[star_BOM+1:end_BOM]
    # print(po_value_BOM,po_index_BOM)

    data_all = []
    idx = 0
    while True:
        data_1 = data
        start_no += 1
        if worksheet.cell(start_no, 1).value is None:
            end_index = start_no
            break

        data_1[4] = (idx+1) * 10                                 # 产品工序标号
        data_1[5] = worksheet.cell(start_no, 4).value            # 工序描述
        efficiency = 100 / worksheet.cell(start_no, 10).value    # 100/标准产能=效率
        data_1[9] = round(efficiency, 2)
        data_1[10] = data_1[9]
        data_1[11] = data_1[9]
        data_1[12] = data_1[9]

        data_all.append(data_1.copy())
        idx += 1

    # print(data_all)
    write2excel(writefile_path, data_all)

    # print(end_index)
    for i in range(end_index, rows + 1):
        v = worksheet.cell(i, 4).value
        if v is None:
            continue
        elif "合计" in str(v):
            break
    for j in [11, 12, 13, 14]:
        loss = worksheet.cell(i, j).value
        if loss is None:
            continue
        if type(loss) != str and loss >= 0:
            break
    f_info.append(loss)
    # print(loss)
    return po_index_BOM, f_info




def BOM(writefile_path, worksheet, worksheet_formula, po_index_BOM, f_info, replace):

    data = [4011, None, None, 1, None, 100000, 0, None, None, 0, "m", 0, None, None, None, None]
    data[1] = f_info[0]
    data[2] = f_info[1]
    data[4] = f_info[2]
    loss = f_info[3]
    data_all = []
    num_replace = 0
    for idx, j in enumerate(po_index_BOM):
        if (idx != 0) and (j != po_index_BOM[idx-1] + 1):
            break

        # data[6] = (idx + 1) * 10                                             # 行项目号
        data[7] = worksheet.cell(j, 4).value                                  # 子件物料号
        # 由公式逆推值
        value = float(worksheet.cell(j, 11).value)
        formula = worksheet_formula.cell(j, 11).value
        # print(formula)
        if loss == 0 or value == formula or re.search(r'[A-Z]', formula) is None:
            data[9] = np.floor(value * 100)
        else:
            ret = re.search(r'[+\-]\d', formula)
            if ret is None:
                data[9] = np.floor(value * (1 - loss) * 100)
            else:
                v = ret.group()
                if v[0] == '+':
                    add = float(v[1:])
                    data[9] = np.floor(((value - add) * (1 - loss) + add) * 100)
                else:
                    sub = float(v[1:])
                    data[9] = np.floor(((value + sub) * (1 - loss) - sub) * 100)
        data_all.append(data.copy())
        for i in [7, 8, 14]:
            vl = worksheet.cell(j, i).value
            if vl is None:
                continue
            vl = str(vl)
            if "不通用" in vl:
                continue
            ret = re.findall(r'[A-Z]{1,3}\d\d+[A-Z]*', vl)
            if ret == []:
                continue
            num_replace += 1
            data_all[-1][12] = num_replace
            data_all[-1][13] = 2
            data_all[-1][14] = 1
            data_all[-1][15] = 100
            value_7 = data_all[-1][7]
            value_7_3 = value_7[-5:]
            for r in ret:
                data_temp = data_all[-1].copy()
                data_temp[14] = data_temp[14] + 1
                data_temp[15] = None
                r_2 = r.ljust(8, '0') + value_7_3
                # print(r_2)
                for i in replace:
                    if r_2 == i[-13:]:
                        data_temp[7] = i
                        break
                else:
                    data_temp[7] = '600-' + r_2
                # print(data_temp[7])
                data_all.append(data_temp.copy())
            break

    for i in range(len(data_all)):
        data_all[i][6] = (i + 1) * 10

    write2excel(writefile_path, data_all)
    return None


if __name__ == '__main__':
    filespath_model  = easygui.diropenbox(msg='请选择 PCC文件、模板文件及“规格料清单.xlsx” 存放文件夹：')
    filespath_model_process = os.path.join(filespath_model, r'工艺路线批量导入模板.xlsx')
    filespath_model_BOM = os.path.join(filespath_model, r'BOM批量导入模板.xlsx')
    filespath_model_replace = os.path.join(filespath_model, r'规格料清单.xlsx')

    try:
        workbook = openpyxl.load_workbook(filespath_model_replace, data_only=True)
        worksheet = workbook.worksheets[0]
        replace = list(worksheet.columns)[3]
        replace_value = [i.value for i in replace]
        # print(replace_value)
    except:
        print("未找到‘规格料清单.xlsx’文件！！！")
        time.sleep(5)
        sys.exit()


    if not os.path.exists(filespath_model_process):
        workbook = openpyxl.Workbook()
        worksheet_new = workbook.create_sheet('Sheet1', 0)
        worksheet_new.append(['工厂', '物料号', '物料描述', '可选工艺路线', '工序编号', '工序描述',
                            '工作中心', '控制码', '基本数量', '机器', '人工', '设备数量', '作业人数'])
        workbook.save(filespath_model_process)
        print("未找到文件:'工艺路线批量导入模板.xlsx'  已生成-->OK")

    if not os.path.exists(filespath_model_BOM):
        workbook = openpyxl.Workbook()
        worksheet_new = workbook.create_sheet('Sheet1', 0)
        worksheet_new.append(['工厂','父件物料号','父件物料描述','BOM用途','可选BOM','基本数量(父件)','行项目号','子件物料号','子件物料描述',
                            '数量(子件）','单位','损耗率','替代项目组','策略','优先级','使用可能性','项目文本1','项目文本2','主材料',
                            '非标损耗率(%)','递归允许','净废品标识','责任工程师','产品类别','终端客户'])
        workbook.save(filespath_model_BOM)
        print("未找到文件:'BOM批量导入模板.xlsx'  已生成-->OK")


    names = os.listdir(filespath_model)
    len_names = len(names)
    for index, filename in enumerate(names):
        filepath = os.path.join(filespath_model, filename)
        e_end = os.path.splitext(filename)[1]

        try:
            if e_end == '.xls':
                print(f'{filename}-->Error-->只能处理.xlsx文件')
                raise ValueError
            if e_end != '.xlsx':
                print(f'{index + 1}/{len_names}', f'{filename}-->Pass')
                continue
            rows, cols, worksheet, worksheet_formula = files_reader(filepath)

            po_index_BOM, f_info = process(filespath_model_process, rows, cols, worksheet)
            BOM(filespath_model_BOM, worksheet, worksheet_formula, po_index_BOM, f_info, replace_value)
            print(f'{index+1}/{len_names}', f'{filename}-->Success')
        except Exception as e:
            # print(e)
            if filename in ['工艺路线批量导入模板.xlsx', 'BOM批量导入模板.xlsx', '规格料清单.xlsx']:
                print(f'{index + 1}/{len_names}', f'{filename}-->Pass')
            else:
                filedir = os.path.join(filespath_model, '操作失败')
                if not os.path.exists(filedir):
                    os.mkdir(filedir)
                filename2 = os.path.join(filedir,filename)
                shutil.copyfile(filepath, filename2)
                print(f'{index+1}/{len_names}', f'{filename}-->Error')