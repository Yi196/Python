import xlrd,xlwt
# 1.2.0以后的xlrd只能打开xls文件  xlwt只能写入xls文件(写xlsx程序不会报错，但最后文件无法直接打开，会报错)
'''对于xlrd和xlwt，行数和列数从0开始，单元格的行和列也从0开始，例如sheet.row_values(2)表示第三行的内容，sheet.cell(1,2).value表示第二行第三列单元格的内容'''

# 获取工作簿对象：
book = xlrd.open_workbook('excel文件名称')
# 获取所有工作表名称：
names = book.sheet_names()  # 结果为列表
# 根据索引获取工作表对象：
sheet = book.sheet_by_index(0)
# 根据名称获取工作表对象：
sheet = book.sheet_by_name('工作表名称')
# 获取工作表行数：
rows = sheet.nrows
# 获取工作表列数：
cols = sheet.ncols
# 获取工作表某一行的内容：
row = sheet.row_values(0)  # 结果为列表
# 获取工作表某一列的内容：
col = sheet.col_values(0)  # 结果为列表
# 获取工作表某一单元格的内容：
cell = sheet.cell_value(0, 0)
call = sheet.cell(0, 0).value

# xlwt写入excel文件
# 创建工作簿：
book = xlwt.Workbook() # 如果写入中文为乱码，可添加参数encoding = 'utf-8'
# 创建工作表：
sheet = book.add_sheet('Sheet1')
# 向单元格写入内容：
sheet.write(0,0,'内容1')
# 保存工作簿：
book.save('excel文件名称')  # 默认保存在py文件相同路径下，如果该路径下有相同文件，会被新创建的文件覆盖，即xlwt不能修改文件。

# 在Excel后追加内容（只能处理xls文件）
from xlutils import copy
def write2excel(path,data):
    workxls = xlrd.open_workbook(path)
    wbook = copy.copy(workxls)
    w_sheet = wbook.get_sheet(0)  # 索引sheet表
    worksheet = workxls.sheet_by_index(0)
    row = worksheet.nrows

    col = 0
    for i in data:
        for j in i:
            w_sheet.write(row, col, j)
            col += 1
        row += 1
        col = 0

    wbook.save(path)



import openpyxl
# openpyxl模块可实现对excel文件的读、写和修改，只能处理xlsx文件，不能处理xls文件
'''对于openpyxl，行数和列数都从1开始，单元格的行和列也从1开始。例如sheet.cell(1,2).value表示第一行第二列单元格的内容'''

# 获取工作簿对象：
book = openpyxl.load_workbook('excel文件名称')
book = openpyxl.load_workbook('excel文件名称', data_only=False)  # data_only=False 表示读取公式
# 获取所有工作表名称：
names = book.sheetnames
# 获取工作表对象：
sheet1 = book.worksheets[0]
sheet2 = book['工作表名称']
sheet3 = book[book.sheetnames[0]]
# 获取工作表名称：
title = sheet1.title
# 获取工作表行数：
rows = sheet1.max_row
# 获取工作表列数：
cols = sheet1.max_column
# 获取某一行列内容
row_1 = list(sheet.rows)[0]
col_4 = list(sheet.columns)[3]
# 获取某一单元格内容：
cell = sheet.cell(1,2).value
cell = sheet['单元格'].value  # 例如sheet['B1'].value 不区分大小写
# 假设有一fruit2.xlsx，除后缀名其他与上述fruit.xls完全一样

# openpyxl写excel文件
# 创建工作簿：
book = openpyxl.Workbook()  # 如果写入中文为乱码，可添加参数encoding = 'utf-8'
# 创建工作表：
sheet = book.create_sheet('工作表名称', 0)  # 0表示创建的工作表在工作薄最前面
# 向单元格写入内容：
sheet.cell(0, 0, '内容1')
# 保存工作簿：
book.save('excel文件名称')  # 默认保存在py文件相同路径下，如果该路径下有相同文件，会被新创建的文件覆盖

# openpyxl修改excel文件
# 在第m行、第n列前面插入行、列
sheet.insert_rows(1)
sheet.insert_cols(2)
# 删除第m行、第n列
sheet.delete_rows(1)
sheet.delete_cols(2)
# 在最后追加行：
sheet.append('可迭代对象')
